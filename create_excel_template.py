import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import os

def create_excel_template():
    """Create a comprehensive Excel template for data collection"""
    
    # Create workbook
    wb = Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    category_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_alignment = Alignment(horizontal='center', vertical='center')
    
    # Task definitions
    tasks = {
        "Factual & Rewriting (1-10)": [
            "Who is the current UN Secretary-General?",
            "Summarise a 150-word news article in one sentence.",
            "Translate a paragraph about climate change into French.",
            "Classify sentiment of 3 tweets (positive/negative).",
            "Extract email & phone number from a text.",
            "Explain the difference between RAM and ROM.",
            "Name three renewable energy sources.",
            "Turn a dense paragraph into bullet points.",
            "Rewrite a sentence in a formal business tone.",
            "Create a catchy blog title about electric cars."
        ],
        "Reasoning & Quantitative (11-15)": [
            "Train A leaves at 10:00 at 100 km/h, Train B at 11:00 at 120 km/h — when do they meet?",
            "Solve a system: 2x+3y=12 and x-y=4.",
            "Give the derivative of x³+2x²-5x+7.",
            "Explain 'overfitting' simply.",
            "Convert 1500 W to kWh for 24 h and to yearly cost at 0.20 €/kWh."
        ],
        "Programming & Debugging (16-20)": [
            "Write Python code reversing a string.",
            "Fix the bug in 'for i in range(5) print(i)'.",
            "Explain what this recursive Python function returns (teacher gives code).",
            "Suggest an optimisation for a slow SQL query (given).",
            "Explain Big-O complexity of binary search."
        ],
        "Knowledge & Reasoning (21-25)": [
            "Compare nuclear vs solar energy (3 pros / 3 cons each).",
            "Explain GDPR compliance steps for a SaaS startup.",
            "Summarise a Wikipedia article on climate change into 5 key bullet points.",
            "Describe in detail the transformer architecture (attention, encoder/decoder).",
            "List and explain three differences between supervised, unsupervised, and reinforcement learning."
        ],
        "Advanced & Creative (26-30)": [
            "Write a project plan for deploying AI to monitor deforestation using satellites.",
            "Draft a LinkedIn post convincing a company to adopt green AI.",
            "Create a short legal disclaimer about data privacy for an AI chatbot.",
            "Imagine and explain a new business model that uses AI to reduce carbon emissions in logistics.",
            "Analyse a research abstract (teacher provides) and rewrite it for a non-technical policymaker."
        ]
    }
    
    # Models to test
    models = [
        ("LLaMA 3.1 8B", "Small"),
        ("Gemma 8B", "Small"),
        ("Mistral Small", "Medium"),
        ("GPT-OSS 20B", "Medium"),
        ("GPT-5", "Large"),
        ("DeepSeek R1", "Large")
    ]
    
    # Create data collection sheet
    ws_data = wb.create_sheet("Data Collection")
    
    # Headers
    headers = [
        "Task ID", "Task Category", "Task Description", "Model", "Model Size",
        "Quality Score (1-5)", "Latency (sec)", "Energy (kWh)", "CO₂ (kg eq.)", 
        "Cost (€)", "Notes"
    ]
    
    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws_data.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = border
    
    # Write data
    row = 2
    task_id = 1
    
    for category, task_list in tasks.items():
        for task_desc in task_list:
            for model_name, model_size in models:
                # Task ID
                ws_data.cell(row=row, column=1, value=task_id).border = border
                # Task Category
                ws_data.cell(row=row, column=2, value=category).border = border
                # Task Description
                ws_data.cell(row=row, column=3, value=task_desc).border = border
                # Model
                ws_data.cell(row=row, column=4, value=model_name).border = border
                # Model Size
                ws_data.cell(row=row, column=5, value=model_size).border = border
                # Quality Score (empty for user input)
                ws_data.cell(row=row, column=6, value="").border = border
                # Latency (empty for user input)
                ws_data.cell(row=row, column=7, value="").border = border
                # Energy (empty for user input)
                ws_data.cell(row=row, column=8, value="").border = border
                # CO₂ (empty for user input)
                ws_data.cell(row=row, column=9, value="").border = border
                # Cost (empty for user input)
                ws_data.cell(row=row, column=10, value="").border = border
                # Notes (empty for user input)
                ws_data.cell(row=row, column=11, value="").border = border
                
                row += 1
            task_id += 1
    
    # Apply category coloring
    current_category = None
    for row in range(2, ws_data.max_row + 1):
        category_cell = ws_data.cell(row=row, column=2)
        if category_cell.value != current_category:
            current_category = category_cell.value
            fill_color = category_fill
        else:
            fill_color = None
        
        for col in range(1, 12):
            cell = ws_data.cell(row=row, column=col)
            if fill_color:
                cell.fill = fill_color
            cell.alignment = center_alignment
    
    # Auto-adjust column widths
    for column in ws_data.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws_data.column_dimensions[column_letter].width = adjusted_width
    
    # Create instructions sheet
    ws_instructions = wb.create_sheet("Instructions")
    
    instructions = [
        ["Compar'IA Benchmarking - Data Collection Instructions"],
        [""],
        ["1. DATA COLLECTION PROCESS"],
        ["   • Test each model on all 30 tasks using the Compar'IA platform"],
        ["   • Record the metrics provided by the platform for each run"],
        ["   • Fill in the 'Data Collection' sheet with your results"],
        [""],
        ["2. METRICS TO RECORD"],
        ["   • Quality Score: Rate the answer quality from 1-5 (1=Poor, 5=Excellent)"],
        ["   • Latency: Response time in seconds"],
        ["   • Energy: Power consumption in kWh"],
        ["   • CO₂: Carbon footprint in kg CO₂ equivalent"],
        ["   • Cost: Financial cost per task in euros"],
        ["   • Notes: Any additional observations"],
        [""],
        ["3. MODELS TO TEST"],
        ["   Small Models:"],
        ["   • LLaMA 3.1 8B"],
        ["   • Gemma 8B"],
        [""],
        ["   Medium Models:"],
        ["   • Mistral Small"],
        ["   • GPT-OSS 20B"],
        [""],
        ["   Large Models:"],
        ["   • GPT-5"],
        ["   • DeepSeek R1"],
        [""],
        ["4. TASK CATEGORIES"],
        ["   • Factual & Rewriting (Tasks 1-10): Basic factual questions"],
        ["   • Reasoning & Quantitative (Tasks 11-15): Math and logic"],
        ["   • Programming & Debugging (Tasks 16-20): Code tasks"],
        ["   • Knowledge & Reasoning (Tasks 21-25): Complex knowledge"],
        ["   • Advanced & Creative (Tasks 26-30): Multi-step creative tasks"],
        [""],
        ["5. QUALITY SCORING GUIDELINES"],
        ["   5 = Excellent: Complete, accurate, well-structured answer"],
        ["   4 = Good: Mostly correct with minor issues"],
        ["   3 = Average: Adequate but with some errors or gaps"],
        ["   2 = Poor: Significant errors or incomplete"],
        ["   1 = Very Poor: Incorrect or irrelevant answer"],
        [""],
        ["6. DASHBOARD USAGE"],
        ["   • Save this file as 'data_collection_results.xlsx'"],
        ["   • Import the data into the Streamlit dashboard"],
        ["   • Use the dashboard to analyze and visualize results"],
        [""],
        ["7. TIPS FOR ACCURATE DATA COLLECTION"],
        ["   • Test models in the same environment/conditions"],
        ["   • Record metrics immediately after each test"],
        ["   • Be consistent in quality scoring across all models"],
        ["   • Note any technical issues or anomalies"],
        [""],
        ["Good luck with your benchmarking! 🚀"]
    ]
    
    for row, instruction in enumerate(instructions, 1):
        cell = ws_instructions.cell(row=row, column=1, value=instruction[0])
        if instruction[0].startswith("Compar'IA") or instruction[0].startswith("1.") or instruction[0].startswith("2.") or instruction[0].startswith("3.") or instruction[0].startswith("4.") or instruction[0].startswith("5.") or instruction[0].startswith("6.") or instruction[0].startswith("7."):
            cell.font = Font(bold=True)
        elif instruction[0].startswith("   •"):
            cell.font = Font(italic=True)
    
    # Auto-adjust column width for instructions
    ws_instructions.column_dimensions['A'].width = 80
    
    # Create summary sheet
    ws_summary = wb.create_sheet("Summary")
    
    summary_data = [
        ["Compar'IA Benchmarking Summary"],
        [""],
        ["Total Tasks:", "30"],
        ["Total Models:", "6"],
        ["Total Test Runs:", "180"],
        [""],
        ["Models by Size:"],
        ["Small (8B parameters):", "2 models"],
        ["Medium (20B parameters):", "2 models"],
        ["Large (70B+ parameters):", "2 models"],
        [""],
        ["Task Distribution:"],
        ["Factual & Rewriting:", "10 tasks"],
        ["Reasoning & Quantitative:", "5 tasks"],
        ["Programming & Debugging:", "5 tasks"],
        ["Knowledge & Reasoning:", "5 tasks"],
        ["Advanced & Creative:", "5 tasks"],
        [""],
        ["Expected Metrics per Model:"],
        ["Quality Scores:", "30 scores (1-5 scale)"],
        ["Latency Measurements:", "30 measurements (seconds)"],
        ["Energy Consumption:", "30 measurements (kWh)"],
        ["CO₂ Emissions:", "30 measurements (kg eq.)"],
        ["Cost Analysis:", "30 measurements (€)"],
        [""],
        ["Data Collection Status:"],
        ["Completed Tasks:", "0/180"],
        ["Completion Rate:", "0%"],
        [""],
        ["Next Steps:"],
        ["1. Begin testing with Compar'IA platform"],
        ["2. Record results in Data Collection sheet"],
        ["3. Import data into Streamlit dashboard"],
        ["4. Analyze results and create visualizations"],
        ["5. Prepare presentation slides"]
    ]
    
    for row, data in enumerate(summary_data, 1):
        cell = ws_summary.cell(row=row, column=1, value=data[0])
        if len(data) > 1:
            ws_summary.cell(row=row, column=2, value=data[1])
        
        if data[0].startswith("Compar'IA") or data[0].endswith(":"):
            cell.font = Font(bold=True)
    
    # Auto-adjust column widths for summary
    ws_summary.column_dimensions['A'].width = 30
    ws_summary.column_dimensions['B'].width = 20
    
    # Save the workbook
    filename = "compar_ia_data_collection_template.xlsx"
    wb.save(filename)
    print(f"✅ Excel template created: {filename}")
    print(f"📊 Total sheets: {len(wb.worksheets)}")
    print(f"📝 Data collection rows: {ws_data.max_row - 1}")
    print(f"🎯 Total test runs: {(ws_data.max_row - 1)}")

if __name__ == "__main__":
    create_excel_template()
