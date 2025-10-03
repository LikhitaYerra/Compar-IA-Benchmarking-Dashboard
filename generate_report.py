#!/usr/bin/env python3
"""
Advanced Analysis Report Generator for ComparAI Benchmarking
Generates comprehensive reports with statistical analysis
"""

import pandas as pd
import numpy as np
from openpyxl import load_workbook
import json
from datetime import datetime

def load_comparai_data():
    """Load data from the ComparAI Excel template"""
    try:
        wb = load_workbook('ComparAI_Benchmark_Template_v2-3.xlsx', data_only=True)
        ws = wb['Runs']
        
        data = []
        headers = []
        
        # Get headers
        for col in range(1, ws.max_column + 1):
            header = ws.cell(row=1, column=col).value
            if header:
                headers.append(str(header))
            else:
                headers.append(f"Column_{col}")
        
        # Get data
        for row in range(2, ws.max_row + 1):
            row_data = []
            has_data = False
            
            for col in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=row, column=col).value
                row_data.append(cell_value)
                if cell_value is not None:
                    has_data = True
            
            if has_data:
                data.append(row_data)
        
        df = pd.DataFrame(data, columns=headers)
        
        # Clean and standardize
        df = clean_comparai_data(df)
        return df
        
    except Exception as e:
        print(f"Error loading data: {e}")
        return None

def clean_comparai_data(df):
    """Clean and standardize the ComparAI data format"""
    df = df.dropna(subset=['Model', 'Task ID'])
    
    column_mapping = {
        'Prompt Category': 'Task_Category',
        'Task ID': 'Task_ID',
        'Task Description': 'Task_Description',
        'Model': 'Model',
        'Quality (1-5)': 'Quality_Score',
        'Latency (sec)': 'Latency_sec',
        'Energy': 'Energy_kWh',
        'co2': 'CO2_kg'
    }
    
    for old_col, new_col in column_mapping.items():
        if old_col in df.columns:
            df[new_col] = df[old_col]
    
    if 'Cost_EUR' not in df.columns:
        df['Cost_EUR'] = 0.0
    
    if 'Notes' not in df.columns:
        df['Notes'] = ''
    
    def get_model_size(model_name):
        if '8B' in str(model_name) or 'Gemma' in str(model_name):
            return 'Small'
        elif 'Small' in str(model_name) or '20B' in str(model_name):
            return 'Medium'
        else:
            return 'Large'
    
    df['Model_Size'] = df['Model'].apply(get_model_size)
    
    numeric_columns = ['Task_ID', 'Quality_Score', 'Latency_sec', 'Energy_kWh', 'CO2_kg', 'Cost_EUR']
    for col in numeric_columns:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='coerce')
    
    df = df.dropna(subset=['Quality_Score', 'Latency_sec', 'Energy_kWh'])
    
    return df

def calculate_advanced_metrics(df):
    """Calculate comprehensive metrics and statistical analysis"""
    metrics = df.groupby(['Model', 'Model_Size']).agg({
        'Quality_Score': ['mean', 'std', 'min', 'max', 'count', 'median'],
        'Latency_sec': ['mean', 'std', 'min', 'max', 'median'],
        'Energy_kWh': ['mean', 'std', 'sum'],
        'CO2_kg': ['mean', 'std', 'sum'],
        'Cost_EUR': ['mean', 'std', 'sum']
    }).round(3)
    
    metrics.columns = ['_'.join(col).strip() for col in metrics.columns]
    metrics = metrics.reset_index()
    
    # Basic efficiency metrics
    metrics['Quality_Efficiency'] = metrics['Quality_Score_mean'] / metrics['Energy_kWh_mean']
    metrics['Cost_Efficiency'] = metrics['Quality_Score_mean'] / metrics['Cost_EUR_mean']
    metrics['Speed_Efficiency'] = metrics['Quality_Score_mean'] / metrics['Latency_sec_mean']
    
    # Advanced consistency metrics
    metrics['Quality_Consistency'] = 1 - (metrics['Quality_Score_std'] / metrics['Quality_Score_mean'])
    metrics['Latency_Consistency'] = 1 - (metrics['Latency_sec_std'] / metrics['Latency_sec_mean'])
    metrics['Energy_Consistency'] = 1 - (metrics['Energy_kWh_std'] / metrics['Energy_kWh_mean'])
    
    # Environmental and cost metrics
    metrics['Environmental_Impact'] = (metrics['CO2_kg_mean'] * 0.7 + metrics['Energy_kWh_mean'] * 0.3)
    metrics['Cost_Per_Quality_Point'] = metrics['Cost_EUR_mean'] / metrics['Quality_Score_mean']
    metrics['Energy_Per_Quality_Point'] = metrics['Energy_kWh_mean'] / metrics['Quality_Score_mean']
    
    # Task completion and reliability
    metrics['Task_Completion_Rate'] = metrics['Quality_Score_count'] / 30
    metrics['Quality_Range'] = metrics['Quality_Score_max'] - metrics['Quality_Score_min']
    metrics['Latency_Range'] = metrics['Latency_sec_max'] - metrics['Latency_sec_min']
    
    # Statistical significance indicators
    metrics['Quality_Reliability'] = metrics['Quality_Score_std'] / metrics['Quality_Score_mean']
    metrics['Latency_Reliability'] = metrics['Latency_sec_std'] / metrics['Latency_sec_mean']
    
    return metrics

def perform_statistical_analysis(df, metrics):
    """Perform comprehensive statistical analysis"""
    analysis = {}
    
    # Overall statistics
    analysis['overall'] = {
        'total_tasks': len(df['Task_ID'].unique()),
        'total_models': len(df['Model'].unique()),
        'total_runs': len(df),
        'date_range': f"{df.get('Date', pd.Series(['Unknown'])).min()} to {df.get('Date', pd.Series(['Unknown'])).max()}"
    }
    
    # Model size analysis
    size_analysis = df.groupby('Model_Size').agg({
        'Quality_Score': 'mean',
        'Latency_sec': 'mean',
        'Energy_kWh': 'mean',
        'CO2_kg': 'mean'
    }).round(3)
    analysis['by_size'] = size_analysis.to_dict()
    
    # Task category analysis
    if 'Task_Category' in df.columns:
        category_analysis = df.groupby('Task_Category').agg({
            'Quality_Score': ['mean', 'std'],
            'Latency_sec': ['mean', 'std'],
            'Energy_kWh': ['mean', 'std']
        }).round(3)
        analysis['by_category'] = category_analysis.to_dict()
    
    # Correlation analysis
    numeric_cols = ['Quality_Score', 'Latency_sec', 'Energy_kWh', 'CO2_kg', 'Cost_EUR']
    correlation_matrix = df[numeric_cols].corr()
    analysis['correlations'] = correlation_matrix.to_dict()
    
    # Performance rankings
    rankings = {}
    for metric in ['Quality_Score_mean', 'Latency_sec_mean', 'Energy_kWh_mean', 'CO2_kg_mean', 'Cost_EUR_mean']:
        if metric in metrics.columns:
            sorted_models = metrics.sort_values(metric, ascending=metric != 'Quality_Score_mean')
            rankings[metric] = sorted_models[['Model', metric]].to_dict('records')
    
    analysis['rankings'] = rankings
    
    # Efficiency analysis
    efficiency_metrics = ['Quality_Efficiency', 'Cost_Efficiency', 'Speed_Efficiency', 'Quality_Consistency']
    efficiency_analysis = {}
    for metric in efficiency_metrics:
        if metric in metrics.columns:
            sorted_models = metrics.sort_values(metric, ascending=False)
            efficiency_analysis[metric] = sorted_models[['Model', metric]].to_dict('records')
    
    analysis['efficiency'] = efficiency_analysis
    
    return analysis

def generate_insights(metrics, analysis):
    """Generate actionable insights and recommendations"""
    insights = []
    
    # Best overall model
    best_overall = metrics.loc[metrics['Quality_Score_mean'].idxmax()]
    insights.append(f"🥇 **Best Overall Quality**: {best_overall['Model']} with {best_overall['Quality_Score_mean']:.2f} average quality score")
    
    # Most efficient models
    most_energy_efficient = metrics.loc[metrics['Quality_Efficiency'].idxmax()]
    insights.append(f"⚡ **Most Energy Efficient**: {most_energy_efficient['Model']} with {most_energy_efficient['Quality_Efficiency']:.2f} quality points per kWh")
    
    most_cost_efficient = metrics.loc[metrics['Cost_Efficiency'].idxmax()]
    insights.append(f"💰 **Most Cost Efficient**: {most_cost_efficient['Model']} with {most_cost_efficient['Cost_Efficiency']:.2f} quality points per euro")
    
    fastest = metrics.loc[metrics['Latency_sec_mean'].idxmin()]
    insights.append(f"🏃 **Fastest Model**: {fastest['Model']} with {fastest['Latency_sec_mean']:.1f}s average latency")
    
    # Environmental impact
    lowest_co2 = metrics.loc[metrics['CO2_kg_mean'].idxmin()]
    insights.append(f"🌍 **Lowest CO₂ Impact**: {lowest_co2['Model']} with {lowest_co2['CO2_kg_mean']:.3f} kg CO₂ per task")
    
    # Consistency analysis
    most_consistent = metrics.loc[metrics['Quality_Consistency'].idxmax()]
    insights.append(f"🎯 **Most Consistent**: {most_consistent['Model']} with {most_consistent['Quality_Consistency']:.3f} consistency score")
    
    # Model size recommendations
    small_models = metrics[metrics['Model_Size'] == 'Small']
    medium_models = metrics[metrics['Model_Size'] == 'Medium']
    large_models = metrics[metrics['Model_Size'] == 'Large']
    
    if not small_models.empty:
        best_small = small_models.loc[small_models['Quality_Score_mean'].idxmax()]
        insights.append(f"📱 **Best Small Model**: {best_small['Model']} - ideal for resource-constrained environments")
    
    if not medium_models.empty:
        best_medium = medium_models.loc[medium_models['Quality_Score_mean'].idxmax()]
        insights.append(f"⚖️ **Best Medium Model**: {best_medium['Model']} - good balance of performance and efficiency")
    
    if not large_models.empty:
        best_large = large_models.loc[large_models['Quality_Score_mean'].idxmax()]
        insights.append(f"🚀 **Best Large Model**: {best_large['Model']} - maximum performance for critical tasks")
    
    return insights

def generate_recommendations(metrics, analysis):
    """Generate specific use-case recommendations"""
    recommendations = []
    
    # Use case recommendations
    recommendations.append("## 🎯 Use Case Recommendations")
    recommendations.append("")
    
    # For different scenarios
    scenarios = [
        {
            'name': 'High-Volume Production',
            'criteria': ['Cost_Efficiency', 'Speed_Efficiency'],
            'description': 'Best for high-volume, cost-sensitive applications'
        },
        {
            'name': 'Environmental Focus',
            'criteria': ['Environmental_Impact', 'Quality_Efficiency'],
            'description': 'Best for environmentally conscious applications'
        },
        {
            'name': 'Quality Critical',
            'criteria': ['Quality_Score_mean', 'Quality_Consistency'],
            'description': 'Best for applications where quality is paramount'
        },
        {
            'name': 'Real-time Applications',
            'criteria': ['Latency_sec_mean', 'Speed_Efficiency'],
            'description': 'Best for real-time or low-latency requirements'
        }
    ]
    
    for scenario in scenarios:
        # Find best model for this scenario
        scenario_metrics = metrics.copy()
        for criterion in scenario['criteria']:
            if criterion in scenario_metrics.columns:
                if 'Latency' in criterion or 'Environmental' in criterion:
                    scenario_metrics[criterion] = scenario_metrics[criterion].rank(ascending=True)
                else:
                    scenario_metrics[criterion] = scenario_metrics[criterion].rank(ascending=False)
        
        # Calculate composite score for scenario
        scenario_metrics['Scenario_Score'] = scenario_metrics[scenario['criteria']].mean(axis=1)
        best_model = scenario_metrics.loc[scenario_metrics['Scenario_Score'].idxmax()]
        
        recommendations.append(f"### {scenario['name']}")
        recommendations.append(f"**Recommended Model**: {best_model['Model']}")
        recommendations.append(f"**Description**: {scenario['description']}")
        recommendations.append("")
    
    return recommendations

def save_analysis_report(metrics, analysis, insights, recommendations):
    """Save comprehensive analysis report"""
    # Convert analysis to JSON-serializable format
    def convert_to_serializable(obj):
        if isinstance(obj, dict):
            return {str(k): convert_to_serializable(v) for k, v in obj.items()}
        elif isinstance(obj, (list, tuple)):
            return [convert_to_serializable(item) for item in obj]
        elif isinstance(obj, (np.integer, np.floating)):
            return float(obj)
        elif isinstance(obj, np.ndarray):
            return obj.tolist()
        elif hasattr(obj, 'to_dict'):
            return obj.to_dict()
        else:
            return str(obj)
    
    report = {
        'timestamp': datetime.now().isoformat(),
        'summary': {
            'total_models': len(metrics),
            'total_tasks': analysis['overall']['total_tasks'],
            'total_runs': analysis['overall']['total_runs']
        },
        'metrics': metrics.to_dict('records'),
        'statistical_analysis': convert_to_serializable(analysis),
        'insights': insights,
        'recommendations': recommendations
    }
    
    # Save as JSON
    with open('comparai_analysis_report.json', 'w') as f:
        json.dump(report, f, indent=2, default=str)
    
    # Save metrics as CSV
    metrics.to_csv('comparai_metrics_detailed.csv', index=False)
    
    print("✅ Analysis report saved:")
    print("   - comparai_analysis_report.json")
    print("   - comparai_metrics_detailed.csv")

def main():
    """Main analysis function"""
    print("🔬 ComparAI Advanced Analysis Report Generator")
    print("=" * 50)
    
    # Load data
    print("📊 Loading ComparAI data...")
    df = load_comparai_data()
    
    if df is None:
        print("❌ Failed to load data")
        return
    
    print(f"✅ Loaded {len(df)} data points")
    
    # Calculate metrics
    print("📈 Calculating advanced metrics...")
    metrics = calculate_advanced_metrics(df)
    print(f"✅ Calculated metrics for {len(metrics)} models")
    
    # Perform statistical analysis
    print("🔍 Performing statistical analysis...")
    analysis = perform_statistical_analysis(df, metrics)
    print("✅ Statistical analysis complete")
    
    # Generate insights
    print("💡 Generating insights...")
    insights = generate_insights(metrics, analysis)
    print(f"✅ Generated {len(insights)} insights")
    
    # Generate recommendations
    print("🎯 Generating recommendations...")
    recommendations = generate_recommendations(metrics, analysis)
    print("✅ Recommendations generated")
    
    # Save report
    print("💾 Saving analysis report...")
    save_analysis_report(metrics, analysis, insights, recommendations)
    
    print("\n🎉 Advanced analysis complete!")
    print("\nKey findings:")
    for insight in insights[:5]:  # Show first 5 insights
        print(f"  {insight}")

if __name__ == "__main__":
    main()
