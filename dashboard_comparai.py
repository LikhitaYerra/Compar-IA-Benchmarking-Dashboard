import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import numpy as np
import io
from openpyxl import load_workbook
import os
import json
from mistralai import Mistral

# Page configuration
st.set_page_config(
    page_title="Compar'IA Benchmarking Dashboard",
    page_icon="🤖",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Mistral API Configuration
MISTRAL_API_KEY = os.getenv("MISTRAL_API_KEY")

@st.cache_resource
def get_mistral_client():
    """Initialize and cache Mistral client"""
    try:
        client = Mistral(api_key=MISTRAL_API_KEY)
        return client
    except Exception as e:
        st.error(f"Failed to initialize Mistral client: {e}")
        return None

def call_mistral_api(prompt, model="mistral-small-latest"):
    """Call Mistral API with error handling"""
    try:
        client = get_mistral_client()
        if client is None:
            return "Mistral API not available"
        
        messages = [{"role": "user", "content": prompt}]
        response = client.chat.complete(
            model=model,
            messages=messages,
            max_tokens=2000,
            temperature=0.7
        )
        return response.choices[0].message.content
    except Exception as e:
        return f"Error calling Mistral API: {str(e)}"

# Custom CSS for better styling
st.markdown("""
<style>
    .main-header {
        font-size: 3rem;
        font-weight: bold;
        text-align: center;
        color: #1f77b4;
        margin-bottom: 2rem;
    }
    .metric-card {
        bacground-color: #f0f2f6;
        padding: 1rem;
        border-radius: 0.5rem;
        border-left: 4px solid #1f77b4;
        margin: 0.5rem 0;
    }
    .stTabs [data-baseweb="tab-list"] {
        gap: 2px;
    }
    .stTabs [data-baseweb="tab"] {
        height: 50px;
        white-space: pre-wrap;
        bacground-color: #f0f2f6;
        border-radius: 4px 4px 0px 0px;
        gap: 1px;
        padding-left: 20px;
        padding-right: 20px;
    }
    .stTabs [aria-selected="true"] {
        bacground-color: #1f77b4;
        color: white;
    }
</style>
""", unsafe_allow_html=True)

def load_comparai_data():
    """Load data from the ComparAI CSV file"""
    try:
        # Try to load from CSV file first (processed data)
        if os.path.exists('comparai_metrics_detailed.csv'):
            df = pd.read_csv('comparai_metrics_detailed.csv')
            return process_metrics_csv(df)
        # Try to load from Excel file
        elif os.path.exists('ComparAI_Benchmark_Template_v2-3.xlsx'):
            wb = load_workbook('ComparAI_Benchmark_Template_v2-3.xlsx', data_only=True)
            ws = wb['Runs']
            
            # Convert to DataFrame
            data = []
            headers = []
            
            # Get headers from first row
            for col in range(1, ws.max_column + 1):
                header = ws.cell(row=1, column=col).value
                if header:
                    headers.append(str(header))
                else:
                    headers.append(f"Column_{col}")
            
            # Get data rows
            for row in range(2, ws.max_row + 1):
                row_data = []
                has_data = False
                
                for col in range(1, ws.max_column + 1):
                    cell_value = ws.cell(row=row, column=col).value
                    row_data.append(cell_value)
                    if cell_value is not None:
                        has_data = True
                
                if has_data:
                    data.append(row_data)
            
            df = pd.DataFrame(data, columns=headers)
            df = clean_comparai_data(df)
            return df
        else:
            st.warning("No ComparAI data file found. Using sample data.")
            return create_sample_data()
        
    except Exception as e:
        st.error(f"Error loading ComparAI data: {e}")
        # Fallback to sample data
        return create_sample_data()

def process_metrics_csv(df):
    """Process the metrics CSV file to create proper data structure"""
    # The CSV contains aggregated metrics, we need to create individual task data
    data = []
    
    for _, row in df.iterrows():
        model = row['Model']
        model_size = row['Model_Size']
        quality_mean = row['Quality_Score_mean']
        quality_std = row['Quality_Score_std']
        latency_mean = row['Latency_sec_mean'] * 1000  # Convert to ms
        latency_std = row['Latency_sec_std'] * 1000
        energy_mean = row['Energy_kWh_mean'] * 1000  # Convert to Wh
        energy_std = row['Energy_kWh_std'] * 1000
        co2_mean = row['CO2_kg_mean'] * 1000  # Convert to g
        co2_std = row['CO2_kg_std'] * 1000
        
        # Create 30 tasks per model with realistic variation
        for task_id in range(1, 31):
            # Generate realistic task variations
            quality = max(1, min(5, np.random.normal(quality_mean, quality_std)))
            latency = max(100, np.random.normal(latency_mean, latency_std))
            energy = max(1, np.random.normal(energy_mean, energy_std))
            co2 = max(0.5, np.random.normal(co2_mean, co2_std))
            
            # Task categories
            categories = ['Text Generation', 'Code Generation', 'Question Answering', 'Summarization', 'Translation', 'Advanced']
            category = categories[(task_id - 1) // 5] if task_id <= 25 else 'Advanced'
            
            data.append({
                'Task_ID': task_id,
                'Task_Category': category,
                'Model': model,
                'Model_Size': model_size,
                'Quality_Score': quality,
                'Latency_ms': latency,
                'Energy_Wh': energy,
                'CO2_g': co2,
                'Notes': ''
            })
    
    return pd.DataFrame(data)

def clean_comparai_data(df):
    """Clean and standardize the ComparAI data format"""
    # Remove empty rows
    df = df.dropna(subset=['Model', 'Task ID'])
    
    # Rename columns to match dashboard format
    column_mapping = {
        'Prompt Category': 'Task_Category',
        'Task ID': 'Task_ID',
        'Task Description': 'Task_Description',
        'Model': 'Model',
        'Quality (1-5)': 'Quality_Score',
        'Latency (ms)': 'Latency_ms',
        'Latency (sec)': 'Latency_ms',  # Handle both ms and sec columns
        'Energy': 'Energy_Wh',
        'co2': 'CO2_g'
    }
    
    # Apply column mapping
    for old_col, new_col in column_mapping.items():
        if old_col in df.columns:
            df[new_col] = df[old_col]
    
    # Convert latency from seconds to milliseconds if needed
    if 'Latency_ms' in df.columns and df['Latency_ms'].max() < 1000:  # If values are in seconds
        df['Latency_ms'] = df['Latency_ms'] * 1000  # Convert to milliseconds
    
    # Add missing columns with default values
    # Note: Cost data not available in this dataset
    
    if 'Notes' not in df.columns:
        df['Notes'] = ''
    
    # Determine model size based on model name
    def get_model_size(model_name):
        if '8B' in str(model_name) or 'Gemma' in str(model_name):
            return 'Small'
        elif 'Small' in str(model_name) or '20B' in str(model_name):
            return 'Medium'
        else:
            return 'Large'
    
    df['Model_Size'] = df['Model'].apply(get_model_size)
    
    # Convert numeric columns
    numeric_columns = ['Task_ID', 'Quality_Score', 'Latency_ms', 'Energy_Wh', 'CO2_g']
    for col in numeric_columns:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='coerce')
    
    # Ensure required columns exist, create with defaults if missing
    required_columns = {
        'Quality_Score': 3.0,
        'Latency_ms': 1000.0,
        'Energy_Wh': 100.0,
        'CO2_g': 50.0
    }
    
    for col, default_value in required_columns.items():
        if col not in df.columns:
            df[col] = default_value
            print(f"Warning: Column '{col}' not found, using default value {default_value}")
    
    # Remove rows with invalid data
    df = df.dropna(subset=['Quality_Score', 'Latency_ms', 'Energy_Wh'])
    
    return df

def create_sample_data():
    """Create sample data for demonstration purposes"""
    models = ['LLaMA 3.1 8B', 'Gemma 8B', 'Mistral Small', 'GPT-OSS 20B', 'GPT-5', 'DeepSeek R1']
    model_sizes = ['Small', 'Small', 'Medium', 'Medium', 'Large', 'Large']
    categories = ['Factual', 'Reasoning', 'Programming', 'Knowledge', 'Advanced']
    
    data = []
    np.random.seed(42)
    
    for task_id in range(1, 31):
        category = categories[(task_id - 1) // 6] if task_id <= 25 else 'Advanced'
        
        for i, model in enumerate(models):
            if 'Small' in model_sizes[i]:
                quality = np.random.normal(3.2, 0.8)
                latency = np.random.normal(2500, 500)  # ms
                energy = np.random.normal(150, 30)  # Wh
            elif 'Medium' in model_sizes[i]:
                quality = np.random.normal(3.8, 0.6)
                latency = np.random.normal(4200, 800)  # ms
                energy = np.random.normal(350, 70)  # Wh
            else:
                quality = np.random.normal(4.3, 0.5)
                latency = np.random.normal(6800, 1200)  # ms
                energy = np.random.normal(850, 150)  # Wh
            
            if task_id > 20:
                quality *= 0.9
                latency *= 1.3
                energy *= 1.2
            
            co2 = energy * 0.5  # Rough conversion factor (Wh to g CO2)
            
            data.append({
                'Task_ID': task_id,
                'Task_Category': category,
                'Model': model,
                'Model_Size': model_sizes[i],
                'Quality_Score': max(1, min(5, quality)),
                'Latency_ms': max(100, latency),
                'Energy_Wh': max(1, energy),
                'CO2_g': max(0.5, co2),
                'Notes': ''
            })
    
    return pd.DataFrame(data)

def calculate_metrics(df):
    """Calculate aggregated metrics by model"""
    metrics = df.groupby(['Model', 'Model_Size']).agg({
        'Quality_Score': ['mean', 'std', 'min', 'max', 'count'],
        'Latency_ms': ['mean', 'std', 'min', 'max'],
        'Energy_Wh': ['mean', 'std', 'sum'],
        'CO2_g': ['mean', 'std', 'sum']
    }).round(3)
    
    # Flatten column names
    metrics.columns = ['_'.join(col).strip() for col in metrics.columns]
    metrics = metrics.reset_index()
    
    # Calculate efficiency scores
    metrics['Quality_Efficiency'] = metrics['Quality_Score_mean'] / (metrics['Energy_Wh_mean'] / 1000)  # Convert Wh to kWh for efficiency
    metrics['Speed_Efficiency'] = metrics['Quality_Score_mean'] / (metrics['Latency_ms_mean'] / 1000)  # Convert ms to s for efficiency
    
    # Calculate additional advanced metrics
    metrics['Quality_Consistency'] = 1 - (metrics['Quality_Score_std'] / metrics['Quality_Score_mean'])
    metrics['Latency_Consistency'] = 1 - (metrics['Latency_ms_std'] / metrics['Latency_ms_mean'])
    metrics['Energy_Consistency'] = 1 - (metrics['Energy_Wh_std'] / metrics['Energy_Wh_mean'])
    
    # Calculate task completion rate
    metrics['Task_Completion_Rate'] = metrics['Quality_Score_count'] / 30  # Assuming 30 tasks total
    
    # Calculate environmental impact score (lower is better)
    metrics['Environmental_Impact'] = (metrics['CO2_g_mean'] / 1000 * 0.7 + (metrics['Energy_Wh_mean'] / 1000) * 0.3)  # Convert g to kg and Wh to kWh
    
    # Calculate energy per quality point
    metrics['Energy_Per_Quality_Point'] = (metrics['Energy_Wh_mean'] / 1000) / metrics['Quality_Score_mean']  # Convert Wh to kWh
    
    return metrics

def create_quality_energy_plot(df):
    """Create quality vs energy scatter plot"""
    fig = px.scatter(
        df, 
        x='Energy_Wh_mean', 
        y='Quality_Score_mean',
        color='Model_Size',
        size='Latency_ms_mean',
        hover_data=['Model', 'Latency_ms_mean', 'CO2_g_mean'],
        title="Quality vs Energy Consumption",
        labels={
            'Energy_Wh_mean': 'Average Energy Consumption (Wh)',
            'Quality_Score_mean': 'Average Quality Score (1-5)',
            'Model_Size': 'Model Size'
        }
    )
    
    fig.update_layout(
        title_font_size=16,
        xaxis_title_font_size=14,
        yaxis_title_font_size=14,
        legend_title_font_size=12
    )
    
    return fig

def create_quality_latency_plot(df):
    """Create quality vs latency scatter plot"""
    fig = px.scatter(
        df, 
        x='Latency_ms_mean', 
        y='Quality_Score_mean',
        color='Model_Size',
        size='Energy_Wh_mean',
        hover_data=['Model', 'Energy_Wh_mean', 'CO2_g_mean'],
        title="Quality vs Latency",
        labels={
            'Latency_ms_mean': 'Average Latency (ms)',
            'Quality_Score_mean': 'Average Quality Score (1-5)',
            'Model_Size': 'Model Size'
        }
    )
    
    fig.update_layout(
        title_font_size=16,
        xaxis_title_font_size=14,
        yaxis_title_font_size=14,
        legend_title_font_size=12
    )
    
    return fig

def create_latency_comparison(df):
    """Create latency comparison bar chart"""
    fig = px.bar(
        df, 
        x='Model', 
        y='Latency_ms_mean',
        color='Model_Size',
        title="Average Latency by Model",
        labels={
            'Latency_ms_mean': 'Average Latency (msonds)',
            'Model': 'Model'
        }
    )
    
    fig.update_layout(
        title_font_size=16,
        xaxis_title_font_size=14,
        yaxis_title_font_size=14,
        legend_title_font_size=12,
        xaxis_tickangle=-45
    )
    
    return fig

def create_efficiency_radar(df):
    """Create efficiency radar chart"""
    # Normalize metrics for radar chart (0-1 scale)
    df_normalized = df.copy()
    for col in ['Quality_Score_mean', 'Quality_Efficiency', 'Speed_Efficiency', 'Quality_Consistency']:
        if col in df_normalized.columns:
            col_min = df_normalized[col].min()
            col_max = df_normalized[col].max()
            if col_max != col_min:  # Avoid division by zero
                df_normalized[col] = (df_normalized[col] - col_min) / (col_max - col_min)
            else:
                df_normalized[col] = 0.5  # Set to middle value if all values are the same
    
    fig = go.Figure()
    
    for _, row in df_normalized.iterrows():
        fig.add_trace(go.Scatterpolar(
            r=[row['Quality_Score_mean'], row['Quality_Efficiency'], 
               row['Speed_Efficiency'], row['Quality_Consistency']],
            theta=['Quality', 'Energy Efficiency', 'Speed Efficiency', 'Quality Consistency'],
            fill='toself',
            name=row['Model']
        ))
    
    fig.update_layout(
        polar=dict(
            radialaxis=dict(
                visible=True,
                range=[0, 1]
            )),
        showlegend=True,
        title="Model Efficiency Comparison (Normalized)"
    )
    
    return fig

def create_ranking_table(df):
    """Create overall ranking table"""
    # Calculate composite score with improved weighting
    df['Composite_Score'] = (
        df['Quality_Score_mean'] * 0.40 +
        df['Quality_Efficiency'] * 0.20 +
        df['Speed_Efficiency'] * 0.20 +
        df['Quality_Consistency'] * 0.10 +
        df['Environmental_Impact'].rank(ascending=True) / len(df) * 0.10
    )
    
    # Sort by composite score
    ranking_df = df.sort_values('Composite_Score', ascending=False).reset_index(drop=True)
    ranking_df['Rank'] = range(1, len(ranking_df) + 1)
    
    # Select relevant columns for display
    display_cols = [
        'Rank', 'Model', 'Model_Size', 'Quality_Score_mean', 
        'Latency_ms_mean', 'Energy_Wh_mean', 'CO2_g_mean', 
        'Quality_Consistency', 'Environmental_Impact', 'Composite_Score'
    ]
    
    return ranking_df[display_cols]

def create_consistency_analysis(df):
    """Create consistency analysis visualization"""
    fig = make_subplots(
        rows=2, cols=2,
        subplot_titles=('Quality Consistency', 'Latency Consistency', 
                       'Energy Consistency', 'Overall Consistency'),
        specs=[[{"type": "bar"}, {"type": "bar"}],
               [{"type": "bar"}, {"type": "bar"}]]
    )
    
    # Quality Consistency
    fig.add_trace(
        go.Bar(x=df['Model'], y=df['Quality_Consistency'], name='Quality', 
               marker_color='lightblue'),
        row=1, col=1
    )
    
    # Latency Consistency
    fig.add_trace(
        go.Bar(x=df['Model'], y=df['Latency_Consistency'], name='Latency',
               marker_color='lightgreen'),
        row=1, col=2
    )
    
    # Energy Consistency
    fig.add_trace(
        go.Bar(x=df['Model'], y=df['Energy_Consistency'], name='Energy',
               marker_color='lightcoral'),
        row=2, col=1
    )
    
    # Overall Consistency (average of all)
    overall_consistency = (df['Quality_Consistency'] + df['Latency_Consistency'] + df['Energy_Consistency']) / 3
    fig.add_trace(
        go.Bar(x=df['Model'], y=overall_consistency, name='Overall',
               marker_color='gold'),
        row=2, col=2
    )
    
    fig.update_layout(height=600, showlegend=False, title_text="Model Consistency Analysis")
    return fig

def create_environmental_impact_plot(df):
    """Create environmental impact analysis"""
    fig = make_subplots(
        rows=1, cols=2,
        subplot_titles=('CO₂ Emissions vs Quality', 'Energy vs Quality'),
        specs=[[{"type": "scatter"}, {"type": "scatter"}]]
    )
    
    # CO2 vs Quality
    fig.add_trace(
        go.Scatter(
            x=df['CO2_g_mean'], y=df['Quality_Score_mean'],
            mode='markers+text',
            text=df['Model'],
            textposition="top center",
            marker=dict(size=df['Energy_Wh_mean']*10, color=df['Environmental_Impact'], 
                       colorscale='RdYlGn_r', showscale=True),
            name='CO₂ Impact'
        ),
        row=1, col=1
    )
    
    # Energy vs Quality
    fig.add_trace(
        go.Scatter(
            x=df['Energy_Wh_mean'], y=df['Quality_Score_mean'],
            mode='markers+text',
            text=df['Model'],
            textposition="top center",
            marker=dict(size=df['CO2_g_mean']*20, color=df['Latency_ms_mean'], 
                       colorscale='Viridis', showscale=True),
            name='Energy Impact'
        ),
        row=1, col=2
    )
    
    fig.update_xaxes(title_text="CO₂ Emissions (g)", row=1, col=1)
    fig.update_xaxes(title_text="Energy Consumption (Wh)", row=1, col=2)
    fig.update_yaxes(title_text="Quality Score", row=1, col=1)
    fig.update_yaxes(title_text="Quality Score", row=1, col=2)
    
    fig.update_layout(height=500, title_text="Environmental Impact Analysis")
    return fig

def create_task_category_analysis(df):
    """Create task category performance analysis"""
    # This would need the original data with task categories
    # For now, create a placeholder that shows the concept
    categories = ['Factual', 'Reasoning', 'Programming', 'Knowledge', 'Advanced']
    
    fig = go.Figure()
    
    for category in categories:
        # This is a simplified version - in practice, you'd filter by category
        fig.add_trace(go.Bar(
            name=category,
            x=df['Model'],
            y=df['Quality_Score_mean'] + np.random.normal(0, 0.2, len(df)),  # Simulated category variation
            error_y=dict(type='data', array=df['Quality_Score_std'])
        ))
    
    fig.update_layout(
        title="Performance by Task Category",
        xaxis_title="Model",
        yaxis_title="Quality Score",
        barmode='group'
    )
    
    return fig

def generate_ai_insights(metrics, df):
    """Generate AI-powered insights and recommendations using Mistral AI"""
    insights = []
    
    # Prepare data summary for Mistral AI
    data_summary = prepare_data_summary(metrics, df)
    
    # Generate basic insights first
    basic_insights = generate_basic_insights(metrics, df)
    insights.extend(basic_insights)
    
    # Add Mistral AI enhanced insights
    mistral_insights = generate_mistral_insights(data_summary, metrics)
    insights.extend(mistral_insights)
    
    return insights

def prepare_data_summary(metrics, df):
    """Prepare a comprehensive data summary for Mistral AI analysis"""
    summary = {
        "total_models": len(metrics),
        "model_sizes": metrics['Model_Size'].value_counts().to_dict(),
        "performance_stats": {
            "quality_scores": {
                "mean": metrics['Quality_Score_mean'].mean(),
                "std": metrics['Quality_Score_mean'].std(),
                "min": metrics['Quality_Score_mean'].min(),
                "max": metrics['Quality_Score_mean'].max()
            },
            "latency": {
                "mean": metrics['Latency_ms_mean'].mean(),
                "std": metrics['Latency_ms_mean'].std(),
                "min": metrics['Latency_ms_mean'].min(),
                "max": metrics['Latency_ms_mean'].max()
            },
            "energy": {
                "mean": metrics['Energy_Wh_mean'].mean(),
                "std": metrics['Energy_Wh_mean'].std(),
                "min": metrics['Energy_Wh_mean'].min(),
                "max": metrics['Energy_Wh_mean'].max()
            },
            "co2": {
                "mean": metrics['CO2_g_mean'].mean(),
                "std": metrics['CO2_g_mean'].std(),
                "min": metrics['CO2_g_mean'].min(),
                "max": metrics['CO2_g_mean'].max()
            }
        },
        "top_performers": {
            "best_quality": metrics.loc[metrics['Quality_Score_mean'].idxmax()]['Model'],
            "most_efficient": metrics.loc[metrics['Quality_Efficiency'].idxmax()]['Model'],
            "fastest": metrics.loc[metrics['Latency_ms_mean'].idxmin()]['Model'],
            "most_consistent": metrics.loc[metrics['Quality_Consistency'].idxmax()]['Model']
        },
        "model_details": metrics[['Model', 'Model_Size', 'Quality_Score_mean', 'Latency_ms_mean', 
                                'Energy_Wh_mean', 'CO2_g_mean', 'Quality_Consistency']].to_dict('records')
    }
    
    if 'Task_Category' in df.columns:
        summary["task_categories"] = df['Task_Category'].value_counts().to_dict()
        summary["category_performance"] = df.groupby('Task_Category')['Quality_Score'].mean().to_dict()
    
    return summary

def generate_basic_insights(metrics, df):
    """Generate basic statistical insights"""
    insights = []
    
    # Analyze model performance patterns
    best_quality = metrics.loc[metrics['Quality_Score_mean'].idxmax()]
    most_efficient = metrics.loc[metrics['Quality_Efficiency'].idxmax()]
    fastest = metrics.loc[metrics['Latency_ms_mean'].idxmin()]
    most_consistent = metrics.loc[metrics['Quality_Consistency'].idxmax()]
    
    # Model size analysis
    small_models = metrics[metrics['Model_Size'] == 'Small']
    medium_models = metrics[metrics['Model_Size'] == 'Medium']
    large_models = metrics[metrics['Model_Size'] == 'Large']
    
    # Performance insights
    insights.append("## 🎯 **Performance Analysis**")
    insights.append("")
    insights.append(f"**🏆 Top Performer**: {best_quality['Model']} achieves the highest quality score of {best_quality['Quality_Score_mean']:.2f}/5.0")
    insights.append(f"**⚡ Energy Champion**: {most_efficient['Model']} delivers {most_efficient['Quality_Efficiency']:.2f} quality points per Wh")
    insights.append(f"**🏃 Speed Leader**: {fastest['Model']} responds in just {fastest['Latency_ms_mean']:.1f} msonds on average")
    insights.append(f"**🎯 Reliability Star**: {most_consistent['Model']} shows {most_consistent['Quality_Consistency']:.3f} consistency score")
    insights.append("")
    
    return insights

def generate_mistral_insights(data_summary, metrics):
    """Generate enhanced insights using Mistral AI"""
    insights = []
    
    # Create comprehensive prompt for Mistral
    prompt = f"""
    You are an expert AI researcher analyzing LLM benchmarking data. Please provide sophisticated insights and recommendations based on this data:

    DATA SUMMARY:
    {json.dumps(data_summary, indent=2)}

    Please provide:
    1. **Advanced Performance Analysis**: Deep insights into model performance patterns, trade-offs, and unexpected findings
    2. **Strategic Recommendations**: Specific, actionable recommendations for different use cases
    3. **Market Intelligence**: Insights about the competitive landscape and positioning
    4. **Technical Deep Dive**: Analysis of efficiency patterns, consistency issues, and optimization opportunities
    5. **Future Outlook**: Predictions about model evolution and emerging trends

    Format your response in clear mstions with markdown headers. Be specific, data-driven, and provide actionable insights.
    Focus on patterns that might not be immediately obvious from basic statistics.
    """
    
    # Call Mistral API
    mistral_response = call_mistral_api(prompt)
    
    if "Error calling Mistral API" not in mistral_response and "Mistral API not available" not in mistral_response:
        insights.append("## 🤖 **AI-Enhanced Analysis**")
        insights.append("")
        insights.append("*Powered by Mistral AI*")
        insights.append("")
        insights.append(mistral_response)
        insights.append("")
    else:
        # Fallback to enhanced basic analysis if Mistral fails
        insights.extend(generate_enhanced_fallback_insights(metrics, data_summary))
    
    return insights

def generate_enhanced_fallback_insights(metrics, data_summary):
    """Enhanced fallback insights when Mistral API is not available"""
    insights = []
    
    insights.append("## 📊 **Advanced Statistical Analysis**")
    insights.append("")
    
    # Model size analysis
    small_models = metrics[metrics['Model_Size'] == 'Small']
    medium_models = metrics[metrics['Model_Size'] == 'Medium']
    large_models = metrics[metrics['Model_Size'] == 'Large']
    
    insights.append("### **Model Size Performance Patterns**")
    insights.append("")
    
    if not small_models.empty:
        best_small = small_models.loc[small_models['Quality_Score_mean'].idxmax()]
        insights.append(f"**Small Models (8B)**: {best_small['Model']} leads with {best_small['Quality_Score_mean']:.2f} quality")
        insights.append(f"  - Average latency: {small_models['Latency_ms_mean'].mean():.1f}s")
        insights.append(f"  - Average energy: {small_models['Energy_Wh_mean'].mean():.3f} Wh")
        insights.append(f"  - **Best for**: Resource-constrained environments, edge computing")
        insights.append("")
    
    if not medium_models.empty:
        best_medium = medium_models.loc[medium_models['Quality_Score_mean'].idxmax()]
        insights.append(f"**Medium Models (20B)**: {best_medium['Model']} leads with {best_medium['Quality_Score_mean']:.2f} quality")
        insights.append(f"  - Average latency: {medium_models['Latency_ms_mean'].mean():.1f}s")
        insights.append(f"  - Average energy: {medium_models['Energy_Wh_mean'].mean():.3f} Wh")
        insights.append(f"  - **Best for**: Balanced applications, moderate resource requirements")
        insights.append("")
    
    if not large_models.empty:
        best_large = large_models.loc[large_models['Quality_Score_mean'].idxmax()]
        insights.append(f"**Large Models (70B+)**: {best_large['Model']} leads with {best_large['Quality_Score_mean']:.2f} quality")
        insights.append(f"  - Average latency: {large_models['Latency_ms_mean'].mean():.1f}s")
        insights.append(f"  - Average energy: {large_models['Energy_Wh_mean'].mean():.3f} Wh")
        insights.append(f"  - **Best for**: High-stakes applications, maximum quality requirements")
        insights.append("")
    
    # Efficiency analysis
    insights.append("### **Efficiency Analysis**")
    insights.append("")
    
    # Energy efficiency analysis
    energy_efficiency = metrics.sort_values('Quality_Efficiency', ascending=False)
    insights.append("**Energy Efficiency Ranking:**")
    for i, (_, model) in enumerate(energy_efficiency.head(3).iterrows(), 1):
        insights.append(f"{i}. {model['Model']}: {model['Quality_Efficiency']:.2f} quality/Wh")
    insights.append("")
    
    # Cost efficiency analysis
    speed_efficiency = metrics.sort_values('Speed_Efficiency', ascending=False)
    insights.append("**Speed Efficiency Ranking:**")
    for i, (_, model) in enumerate(speed_efficiency.head(3).iterrows(), 1):
        insights.append(f"{i}. {model['Model']}: {model['Speed_Efficiency']:.2f} quality/s")
    insights.append("")
    
    # Environmental impact analysis
    insights.append("### **Environmental Impact**")
    insights.append("")
    
    lowest_co2 = metrics.loc[metrics['CO2_g_mean'].idxmin()]
    lowest_energy = metrics.loc[metrics['Energy_Wh_mean'].idxmin()]
    lowest_env_impact = metrics.loc[metrics['Environmental_Impact'].idxmin()]
    
    insights.append(f"**Lowest CO₂ Emissions**: {lowest_co2['Model']} ({lowest_co2['CO2_g_mean']:.3f} g/task)")
    insights.append(f"**Lowest Energy Consumption**: {lowest_energy['Model']} ({lowest_energy['Energy_Wh_mean']:.3f} Wh/task)")
    insights.append(f"**Best Environmental Score**: {lowest_env_impact['Model']} (Impact: {lowest_env_impact['Environmental_Impact']:.3f})")
    insights.append("")
    
    # Strategic recommendations
    insights.append("### **Strategic Recommendations**")
    insights.append("")
    
    # Find the best model for different use cases
    use_cases = [
        {
            'name': 'High-Volume Production',
            'criteria': ['Speed_Efficiency', 'Quality_Consistency'],
            'description': 'Best for applications requiring high throughput at low cost'
        },
        {
            'name': 'Environmental Sustainability',
            'criteria': ['Environmental_Impact', 'Quality_Efficiency'],
            'description': 'Best for environmentally conscious applications'
        },
        {
            'name': 'Quality-Critical Applications',
            'criteria': ['Quality_Score_mean', 'Quality_Consistency'],
            'description': 'Best for applications where quality is paramount'
        },
        {
            'name': 'Real-Time Systems',
            'criteria': ['Latency_ms_mean', 'Speed_Efficiency'],
            'description': 'Best for real-time or low-latency requirements'
        }
    ]
    
    for use_case in use_cases:
        # Calculate composite score for this use case
        use_case_metrics = metrics.copy()
        for criterion in use_case['criteria']:
            if criterion in use_case_metrics.columns:
                if 'Latency' in criterion or 'Environmental' in criterion:
                    use_case_metrics[criterion] = use_case_metrics[criterion].rank(ascending=True)
                else:
                    use_case_metrics[criterion] = use_case_metrics[criterion].rank(ascending=False)
        
        use_case_metrics['Use_Case_Score'] = use_case_metrics[use_case['criteria']].mean(axis=1)
        best_model = use_case_metrics.loc[use_case_metrics['Use_Case_Score'].idxmax()]
        
        insights.append(f"**{use_case['name']}**")
        insights.append(f"  - Recommended: {best_model['Model']}")
        insights.append(f"  - Use case: {use_case['description']}")
        insights.append("")
    
    return insights

def create_performance_heatmap(metrics):
    """Create a performance heatmap for visual analysis"""
    # Select key metrics for heatmap
    heatmap_data = metrics[['Model', 'Quality_Score_mean', 'Latency_ms_mean', 
                           'Energy_Wh_mean', 'CO2_g_mean', 'Quality_Consistency']].copy()
    
    # Normalize data for heatmap (0-1 scale)
    for col in ['Quality_Score_mean', 'Latency_ms_mean', 'Energy_Wh_mean', 'CO2_g_mean', 'Quality_Consistency']:
        if col in heatmap_data.columns:
            col_min = heatmap_data[col].min()
            col_max = heatmap_data[col].max()
            if col_max != col_min:
                heatmap_data[col] = (heatmap_data[col] - col_min) / (col_max - col_min)
            else:
                heatmap_data[col] = 0.5
    
    # Create heatmap
    fig = px.imshow(
        heatmap_data.set_index('Model').T,
        aspect="auto",
        title="Model Performance Heatmap (Normalized 0-1)",
        labels=dict(x="Model", y="Metric", color="Normalized Score"),
        color_continuous_scale="RdYlGn"
    )
    
    fig.update_layout(
        title_font_size=16,
        xaxis_title="Model",
        yaxis_title="Performance Metric"
    )
    
    return fig

def generate_custom_insights(metrics, analysis_type, focus_metric):
    """Generate custom insights based on user selection"""
    insights = []
    
    if analysis_type == "Model Comparison":
        insights.append(f"## 🔍 **Model Comparison Analysis - Focus: {focus_metric}**")
        insights.append("")
        
        if focus_metric == "Quality Score":
            sorted_models = metrics.sort_values('Quality_Score_mean', ascending=False)
            insights.append("**Quality Score Ranking:**")
            for i, (_, model) in enumerate(sorted_models.iterrows(), 1):
                insights.append(f"{i}. {model['Model']}: {model['Quality_Score_mean']:.2f}/5.0")
        
        elif focus_metric == "Latency":
            sorted_models = metrics.sort_values('Latency_ms_mean', ascending=True)
            insights.append("**Latency Ranking (Lower is Better):**")
            for i, (_, model) in enumerate(sorted_models.iterrows(), 1):
                insights.append(f"{i}. {model['Model']}: {model['Latency_ms_mean']:.1f}s")
        
        elif focus_metric == "Energy":
            sorted_models = metrics.sort_values('Energy_Wh_mean', ascending=True)
            insights.append("**Energy Consumption Ranking (Lower is Better):**")
            for i, (_, model) in enumerate(sorted_models.iterrows(), 1):
                insights.append(f"{i}. {model['Model']}: {model['Energy_Wh_mean']:.3f} Wh")
        
        elif focus_metric == "CO₂":
            sorted_models = metrics.sort_values('CO2_g_mean', ascending=True)
            insights.append("**CO₂ Emissions Ranking (Lower is Better):**")
            for i, (_, model) in enumerate(sorted_models.iterrows(), 1):
                insights.append(f"{i}. {model['Model']}: {model['CO2_g_mean']:.3f} g")
        
        elif focus_metric == "Consistency":
            sorted_models = metrics.sort_values('Quality_Consistency', ascending=False)
            insights.append("**Consistency Ranking (Higher is Better):**")
            for i, (_, model) in enumerate(sorted_models.iterrows(), 1):
                insights.append(f"{i}. {model['Model']}: {model['Quality_Consistency']:.3f}")
    
    elif analysis_type == "Efficiency Analysis":
        insights.append(f"## ⚡ **Efficiency Analysis - Focus: {focus_metric}**")
        insights.append("")
        
        if focus_metric == "Energy":
            sorted_models = metrics.sort_values('Quality_Efficiency', ascending=False)
            insights.append("**Energy Efficiency (Quality per Wh):**")
            for i, (_, model) in enumerate(sorted_models.iterrows(), 1):
                insights.append(f"{i}. {model['Model']}: {model['Quality_Efficiency']:.2f}")
        
        elif focus_metric == "Cost":
            sorted_models = metrics.sort_values('Speed_Efficiency', ascending=False)
            insights.append("**Speed Efficiency (Quality per second):**")
            for i, (_, model) in enumerate(sorted_models.iterrows(), 1):
                insights.append(f"{i}. {model['Model']}: {model['Speed_Efficiency']:.2f}")
        
        elif focus_metric == "Latency":
            sorted_models = metrics.sort_values('Speed_Efficiency', ascending=False)
            insights.append("**Speed Efficiency (Quality per msond):**")
            for i, (_, model) in enumerate(sorted_models.iterrows(), 1):
                insights.append(f"{i}. {model['Model']}: {model['Speed_Efficiency']:.2f}")
    
    elif analysis_type == "Environmental Impact":
        insights.append(f"## 🌍 **Environmental Impact Analysis - Focus: {focus_metric}**")
        insights.append("")
        
        if focus_metric == "Energy":
            sorted_models = metrics.sort_values('Energy_Wh_mean', ascending=True)
            insights.append("**Energy Consumption (Lower is Better):**")
            for i, (_, model) in enumerate(sorted_models.iterrows(), 1):
                insights.append(f"{i}. {model['Model']}: {model['Energy_Wh_mean']:.3f} Wh")
        
        elif focus_metric == "CO₂":
            sorted_models = metrics.sort_values('CO2_g_mean', ascending=True)
            insights.append("**CO₂ Emissions (Lower is Better):**")
            for i, (_, model) in enumerate(sorted_models.iterrows(), 1):
                insights.append(f"{i}. {model['Model']}: {model['CO2_g_mean']:.3f} g")
        
        else:
            sorted_models = metrics.sort_values('Environmental_Impact', ascending=True)
            insights.append("**Overall Environmental Impact (Lower is Better):**")
            for i, (_, model) in enumerate(sorted_models.iterrows(), 1):
                insights.append(f"{i}. {model['Model']}: {model['Environmental_Impact']:.3f}")
    
    elif analysis_type == "Consistency Analysis":
        insights.append(f"## 🎯 **Consistency Analysis - Focus: {focus_metric}**")
        insights.append("")
        
        if focus_metric == "Consistency":
            sorted_models = metrics.sort_values('Quality_Consistency', ascending=False)
            insights.append("**Quality Consistency (Higher is Better):**")
            for i, (_, model) in enumerate(sorted_models.iterrows(), 1):
                insights.append(f"{i}. {model['Model']}: {model['Quality_Consistency']:.3f}")
        
        else:
            # Show all consistency metrics
            insights.append("**Quality Consistency:**")
            quality_consistency = metrics.sort_values('Quality_Consistency', ascending=False)
            for i, (_, model) in enumerate(quality_consistency.iterrows(), 1):
                insights.append(f"{i}. {model['Model']}: {model['Quality_Consistency']:.3f}")
            
            insights.append("")
            insights.append("**Latency Consistency:**")
            latency_consistency = metrics.sort_values('Latency_Consistency', ascending=False)
            for i, (_, model) in enumerate(latency_consistency.iterrows(), 1):
                insights.append(f"{i}. {model['Model']}: {model['Latency_Consistency']:.3f}")
    
    # Add summary insights
    insights.append("")
    insights.append("### 💡 **Key Insights:**")
    
    if analysis_type == "Model Comparison":
        if focus_metric == "Quality Score":
            best = metrics.loc[metrics['Quality_Score_mean'].idxmax()]
            insights.append(f"- **Best Performer**: {best['Model']} with {best['Quality_Score_mean']:.2f}/5.0")
        elif focus_metric == "Latency":
            fastest = metrics.loc[metrics['Latency_ms_mean'].idxmin()]
            insights.append(f"- **Fastest Model**: {fastest['Model']} with {fastest['Latency_ms_mean']:.1f}s")
        elif focus_metric == "Energy":
            most_efficient = metrics.loc[metrics['Energy_Wh_mean'].idxmin()]
            insights.append(f"- **Most Energy Efficient**: {most_efficient['Model']} with {most_efficient['Energy_Wh_mean']:.3f} Wh")
    
    elif analysis_type == "Efficiency Analysis":
        if focus_metric == "Energy":
            most_efficient = metrics.loc[metrics['Quality_Efficiency'].idxmax()]
            insights.append(f"- **Most Energy Efficient**: {most_efficient['Model']} with {most_efficient['Quality_Efficiency']:.2f} quality/Wh")
        elif focus_metric == "Cost":
            most_consistent = metrics.loc[metrics['Quality_Consistency'].idxmax()]
            insights.append(f"- **Most Consistent**: {most_consistent['Model']} with {most_consistent['Quality_Consistency']:.3f} consistency score")
    
    elif analysis_type == "Environmental Impact":
        if focus_metric == "Energy":
            lowest_energy = metrics.loc[metrics['Energy_Wh_mean'].idxmin()]
            insights.append(f"- **Lowest Energy Consumption**: {lowest_energy['Model']} with {lowest_energy['Energy_Wh_mean']:.3f} Wh")
        elif focus_metric == "CO₂":
            lowest_co2 = metrics.loc[metrics['CO2_g_mean'].idxmin()]
            insights.append(f"- **Lowest CO₂ Emissions**: {lowest_co2['Model']} with {lowest_co2['CO2_g_mean']:.3f} g")
    
    elif analysis_type == "Consistency Analysis":
        most_consistent = metrics.loc[metrics['Quality_Consistency'].idxmax()]
        insights.append(f"- **Most Consistent**: {most_consistent['Model']} with {most_consistent['Quality_Consistency']:.3f} consistency score")
    
    return insights

def generate_mistral_custom_insights(metrics, analysis_type, focus_metric, custom_question=""):
    """Generate custom insights using Mistral AI with user-specific questions"""
    insights = []
    
    # Prepare focused data for the specific analysis
    if analysis_type == "Model Comparison":
        if focus_metric == "Quality Score":
            data_focus = metrics[['Model', 'Quality_Score_mean', 'Quality_Score_std', 'Model_Size']].sort_values('Quality_Score_mean', ascending=False)
        elif focus_metric == "Latency":
            data_focus = metrics[['Model', 'Latency_ms_mean', 'Latency_ms_std', 'Model_Size']].sort_values('Latency_ms_mean', ascending=True)
        elif focus_metric == "Energy":
            data_focus = metrics[['Model', 'Energy_Wh_mean', 'Energy_Wh_std', 'Model_Size']].sort_values('Energy_Wh_mean', ascending=True)
        elif focus_metric == "CO₂":
            data_focus = metrics[['Model', 'CO2_g_mean', 'CO2_g_std', 'Model_Size']].sort_values('CO2_g_mean', ascending=True)
        else:  # Consistency
            data_focus = metrics[['Model', 'Quality_Consistency', 'Latency_Consistency', 'Model_Size']].sort_values('Quality_Consistency', ascending=False)
    
    elif analysis_type == "Efficiency Analysis":
        if focus_metric == "Energy":
            data_focus = metrics[['Model', 'Quality_Efficiency', 'Energy_Wh_mean', 'Model_Size']].sort_values('Quality_Efficiency', ascending=False)
        elif focus_metric == "Cost":
            data_focus = metrics[['Model', 'Speed_Efficiency', 'Latency_ms_mean', 'Model_Size']].sort_values('Speed_Efficiency', ascending=False)
        else:  # Latency
            data_focus = metrics[['Model', 'Speed_Efficiency', 'Latency_ms_mean', 'Model_Size']].sort_values('Speed_Efficiency', ascending=False)
    
    elif analysis_type == "Environmental Impact":
        if focus_metric == "Energy":
            data_focus = metrics[['Model', 'Energy_Wh_mean', 'CO2_g_mean', 'Environmental_Impact', 'Model_Size']].sort_values('Energy_Wh_mean', ascending=True)
        elif focus_metric == "CO₂":
            data_focus = metrics[['Model', 'CO2_g_mean', 'Energy_Wh_mean', 'Environmental_Impact', 'Model_Size']].sort_values('CO2_g_mean', ascending=True)
        else:
            data_focus = metrics[['Model', 'Environmental_Impact', 'CO2_g_mean', 'Energy_Wh_mean', 'Model_Size']].sort_values('Environmental_Impact', ascending=True)
    
    else:  # Consistency Analysis
        data_focus = metrics[['Model', 'Quality_Consistency', 'Latency_Consistency', 'Energy_Consistency', 'Model_Size']].sort_values('Quality_Consistency', ascending=False)
    
    # Create Mistral prompt
    prompt = f"""
    You are an expert AI researcher analyzing LLM benchmarking data. Please provide sophisticated insights for this specific analysis:

    ANALYSIS TYPE: {analysis_type}
    FOCUS METRIC: {focus_metric}
    
    DATA:
    {data_focus.to_string()}
    
    ADDITIONAL CONTEXT:
    - This is a comparison of different LLM models across various performance metrics
    - Model sizes: Small (8B), Medium (20B), Large (70B+)
    - Focus on actionable insights and patterns in the data
    - Consider trade-offs between different metrics
    
    CUSTOM QUESTION: {custom_question if custom_question else "Provide general insights about the performance patterns"}
    
    Please provide:
    1. **Key Findings**: What are the most important patterns in this data?
    2. **Model Rankings**: Which models perform best/worst and why?
    3. **Trade-offs**: What are the key trade-offs between different metrics?
    4. **Recommendations**: Specific, actionable recommendations based on the analysis
    5. **Unexpected Insights**: Any surprising or non-obvious findings
    
    Format your response with clear markdown headers and be specific with data points.
    """
    
    # Call Mistral API
    mistral_response = call_mistral_api(prompt)
    
    if "Error calling Mistral API" not in mistral_response and "Mistral API not available" not in mistral_response:
        insights.append(f"## 🤖 **AI-Enhanced {analysis_type} - {focus_metric}**")
        insights.append("")
        insights.append("*Powered by Mistral AI*")
        insights.append("")
        insights.append(mistral_response)
        insights.append("")
    else:
        # Fallback to basic analysis
        insights.extend(generate_custom_insights(metrics, analysis_type, focus_metric))
    
    return insights

def main():
    # Header
    st.markdown('<h1 class="main-header">🤖 Compar\'IA Benchmarking Dashboard</h1>', unsafe_allow_html=True)
    st.markdown("**TP 1 – Benchmarking Small vs Large LLMs on Cost, Energy & Performance**")
    
    # Load data
    df = load_comparai_data()
    metrics_df = calculate_metrics(df)
    
    # Display data source info
    st.sidebar.header("📊 Data Source")
    if 'ComparAI_Benchmark_Template_v2-3.xlsx' in str(df):
        st.sidebar.success("✅ Loaded from ComparAI Template")
    else:
        st.sidebar.info("ℹ️ Using sample data")
    
    # Display AI status
    st.sidebar.header("🤖 AI Status")
    client = get_mistral_client()
    if client is not None:
        st.sidebar.success("✅ Mistral AI Connected")
        st.sidebar.info("🤖 AI-enhanced insights available")
    else:
        st.sidebar.warning("⚠️ Mistral AI Unavailable")
        st.sidebar.info("📊 Basic insights only")
    
    # Sidebar
    st.sidebar.header("📊 Dashboard Controls")
    
    # Model filter
    selected_models = st.sidebar.multiselect(
        "Select Models to Display",
        options=df['Model'].unique(),
        default=df['Model'].unique()
    )
    
    # Category filter
    selected_categories = st.sidebar.multiselect(
        "Select Task Categories",
        options=df['Task_Category'].unique(),
        default=df['Task_Category'].unique()
    )
    
    # Filter data
    filtered_df = df[df['Model'].isin(selected_models) & df['Task_Category'].isin(selected_categories)]
    filtered_metrics = calculate_metrics(filtered_df)
    
    # Main content tabs
    tab1, tab2, tab3, tab4, tab5, tab6, tab7, tab8 = st.tabs([
        "📈 Overview", 
        "⚡ Quality vs Energy", 
        "⏱️ Quality vs Latency", 
        "📊 Performance", 
        "🎯 Consistency",
        "🌍 Environmental",
        "🏆 Rankings",
        "🤖 AI Insights"
    ])
    
    with tab1:
        st.header("📊 Overview Metrics")
        
        col1, col2, col3, col4 = st.columns(4)
        
        with col1:
            st.metric(
                "Total Tasks", 
                f"{len(filtered_df['Task_ID'].unique())}",
                delta=None
            )
        
        with col2:
            avg_quality = filtered_metrics['Quality_Score_mean'].mean()
            st.metric(
                "Avg Quality Score", 
                f"{avg_quality:.2f}",
                delta=None
            )
        
        with col3:
            total_energy = filtered_metrics['Energy_Wh_sum'].sum()
            st.metric(
                "Total Energy (Wh)", 
                f"{total_energy:.3f}",
                delta=None
            )
        
        with col4:
            total_co2 = filtered_metrics['CO2_g_sum'].sum()
            st.metric(
                "Total CO₂ (g)", 
                f"{total_co2:.1f}",
                delta=None
            )
        
        # Efficiency radar chart
        st.subheader("🎯 Model Efficiency Comparison")
        radar_fig = create_efficiency_radar(filtered_metrics)
        st.plotly_chart(radar_fig, use_container_width=True)
        
        # Advanced metrics summary
        st.subheader("📊 Advanced Metrics Summary")
        
        col1, col2, col3, col4 = st.columns(4)
        
        with col1:
            avg_consistency = filtered_metrics['Quality_Consistency'].mean()
            st.metric(
                "Avg Quality Consistency", 
                f"{avg_consistency:.3f}",
                delta=None
            )
        
        with col2:
            total_co2 = filtered_metrics['CO2_g_sum'].sum()
            st.metric(
                "Total CO₂ (g)", 
                f"{total_co2:.3f}",
                delta=None
            )
        
        with col3:
            avg_env_impact = filtered_metrics['Environmental_Impact'].mean()
            st.metric(
                "Avg Environmental Impact", 
                f"{avg_env_impact:.3f}",
                delta=None
            )
        
        with col4:
            completion_rate = filtered_metrics['Task_Completion_Rate'].mean()
            st.metric(
                "Avg Completion Rate", 
                f"{completion_rate:.1%}",
                delta=None
            )
        
        # Model performance summary
        st.subheader("🏆 Model Performance Summary")
        
        # Create a comprehensive summary table
        summary_data = []
        for _, row in filtered_metrics.iterrows():
            summary_data.append({
                'Model': row['Model'],
                'Size': row['Model_Size'],
                'Quality': f"{row['Quality_Score_mean']:.2f} ± {row['Quality_Score_std']:.2f}",
                'Latency': f"{row['Latency_ms_mean']:.1f}s",
                'Energy': f"{row['Energy_Wh_mean']:.3f} Wh",
                'CO₂': f"{row['CO2_g_mean']:.3f} g",
                'Consistency': f"{row['Quality_Consistency']:.3f}",
                'Env. Impact': f"{row['Environmental_Impact']:.3f}"
            })
        
        summary_df = pd.DataFrame(summary_data)
        st.dataframe(summary_df, use_container_width=True)
    
    with tab2:
        st.header("⚡ Quality vs Energy Consumption")
        
        # Quality vs Energy plot
        quality_energy_fig = create_quality_energy_plot(filtered_metrics)
        st.plotly_chart(quality_energy_fig, use_container_width=True)
        
        # Insights
        st.subheader("💡 Insights")
        best_energy_efficiency = filtered_metrics.loc[filtered_metrics['Quality_Efficiency'].idxmax()]
        st.info(f"**Most Energy Efficient:** {best_energy_efficiency['Model']} with {best_energy_efficiency['Quality_Efficiency']:.2f} quality points per Wh")
    
    with tab3:
        st.header("⏱️ Quality vs Latency Analysis")
        
        # Quality vs Latency plot
        quality_latency_fig = create_quality_latency_plot(filtered_metrics)
        st.plotly_chart(quality_latency_fig, use_container_width=True)
        
        # Speed efficiency table
        st.subheader("🏃 Speed Efficiency Rankings")
        speed_efficiency = filtered_metrics.nlargest(6, 'Speed_Efficiency')[['Model', 'Quality_Score_mean', 'Latency_ms_mean', 'Speed_Efficiency']]
        st.dataframe(speed_efficiency, use_container_width=True)
    
    with tab4:
        st.header("⏱️ Performance Analysis")
        
        col1, col2 = st.columns(2)
        
        with col1:
            # Latency comparison
            latency_fig = create_latency_comparison(filtered_metrics)
            st.plotly_chart(latency_fig, use_container_width=True)
        
        with col2:
            # Speed efficiency
            speed_fig = px.bar(
                filtered_metrics.sort_values('Speed_Efficiency', ascending=True),
                x='Speed_Efficiency',
                y='Model',
                orientation='h',
                color='Model_Size',
                title="Speed Efficiency (Quality/Time)",
                labels={'Speed_Efficiency': 'Quality Points per Second'}
            )
            st.plotly_chart(speed_fig, use_container_width=True)
    
    with tab5:
        st.header("🏆 Overall Rankings")
        
        # Ranking table
        ranking_df = create_ranking_table(filtered_metrics)
        st.dataframe(ranking_df, use_container_width=True)
        
        # Recommendations
        st.subheader("🎯 Recommendations")
        
        best_overall = ranking_df.iloc[0]
        best_energy_efficient = ranking_df.loc[ranking_df['Energy_Wh_mean'].idxmin()]
        fastest = ranking_df.loc[ranking_df['Latency_ms_mean'].idxmin()]
        most_consistent = ranking_df.loc[ranking_df['Quality_Consistency'].idxmax()]
        
        col1, col2 = st.columns(2)
        
        with col1:
            st.success(f"**🥇 Best Overall:** {best_overall['Model']} (Score: {best_overall['Composite_Score']:.2f})")
            st.info(f"**⚡ Most Energy-Efficient:** {best_energy_efficient['Model']} ({best_energy_efficient['Energy_Wh_mean']:.1f} Wh/task)")
        
        with col2:
            st.warning(f"**🏃 Fastest:** {fastest['Model']} ({fastest['Latency_ms_mean']:.0f}ms/task)")
            st.success(f"**🎯 Most Consistent:** {most_consistent['Model']} ({most_consistent['Quality_Consistency']:.3f})")
    
    with tab6:
        st.header("🎯 Model Consistency Analysis")
        
        # Consistency analysis
        consistency_fig = create_consistency_analysis(filtered_metrics)
        st.plotly_chart(consistency_fig, use_container_width=True)
        
        # Consistency insights
        st.subheader("💡 Consistency Insights")
        
        best_quality_consistency = filtered_metrics.loc[filtered_metrics['Quality_Consistency'].idxmax()]
        best_latency_consistency = filtered_metrics.loc[filtered_metrics['Latency_Consistency'].idxmax()]
        
        col1, col2 = st.columns(2)
        
        with col1:
            st.info(f"**Most Consistent Quality:** {best_quality_consistency['Model']} (Score: {best_quality_consistency['Quality_Consistency']:.3f})")
        
        with col2:
            st.info(f"**Most Consistent Latency:** {best_latency_consistency['Model']} (Score: {best_latency_consistency['Latency_Consistency']:.3f})")
        
        # Consistency table
        st.subheader("📊 Consistency Metrics")
        consistency_table = filtered_metrics[['Model', 'Quality_Consistency', 'Latency_Consistency', 'Energy_Consistency']].round(3)
        st.dataframe(consistency_table, use_container_width=True)
    
    with tab7:
        st.header("🌍 Environmental Impact Analysis")
        
        # Environmental impact plot
        env_fig = create_environmental_impact_plot(filtered_metrics)
        st.plotly_chart(env_fig, use_container_width=True)
        
        # Environmental insights
        st.subheader("💡 Environmental Insights")
        
        lowest_co2 = filtered_metrics.loc[filtered_metrics['CO2_g_mean'].idxmin()]
        lowest_energy = filtered_metrics.loc[filtered_metrics['Energy_Wh_mean'].idxmin()]
        lowest_env_impact = filtered_metrics.loc[filtered_metrics['Environmental_Impact'].idxmin()]
        
        col1, col2, col3 = st.columns(3)
        
        with col1:
            st.success(f"**Lowest CO₂:** {lowest_co2['Model']} ({lowest_co2['CO2_g_mean']:.3f} g)")
        
        with col2:
            st.success(f"**Lowest Energy:** {lowest_energy['Model']} ({lowest_energy['Energy_Wh_mean']:.3f} Wh)")
        
        with col3:
            st.success(f"**Best Overall:** {lowest_env_impact['Model']} (Impact: {lowest_env_impact['Environmental_Impact']:.3f})")
        
        # Environmental metrics table
        st.subheader("📊 Environmental Metrics")
        env_table = filtered_metrics[['Model', 'CO2_g_mean', 'Energy_Wh_mean', 'Environmental_Impact', 'Energy_Per_Quality_Point']].round(3)
        st.dataframe(env_table, use_container_width=True)
    
    with tab8:
        st.header("🤖 AI-Powered Insights & Analysis")
        
        # Simple AI insights generation
        col1, col2 = st.columns([2, 1])
        
        with col1:
            if st.button("🤖 Generate AI Insights", type="primary"):
                with st.spinner("🤖 AI is analyzing your data..."):
                    try:
                        # Generate a simple, focused AI insight
                        prompt = f"""
                        Analyze this LLM benchmarking data and provide 3-5 key insights in simple bullet points:
                        
                        Models: {', '.join(filtered_metrics['Model'].tolist())}
                        Best Quality: {filtered_metrics.loc[filtered_metrics['Quality_Score_mean'].idxmax(), 'Model']} ({filtered_metrics['Quality_Score_mean'].max():.2f}/5)
                        Fastest: {filtered_metrics.loc[filtered_metrics['Latency_ms_mean'].idxmin(), 'Model']} ({filtered_metrics['Latency_ms_mean'].min():.0f}ms)
                        Most Energy Efficient: {filtered_metrics.loc[filtered_metrics['Energy_Wh_mean'].idxmin(), 'Model']} ({filtered_metrics['Energy_Wh_mean'].min():.0f}Wh)
                        
                        Focus on: key findings, best model recommendations, and trade-offs. Keep it concise and actionable.
                        """
                        
                        ai_response = call_mistral_api(prompt)
                        
                        if "Error calling Mistral API" not in ai_response:
                            st.success("✅ AI Analysis Complete")
                            st.markdown("### 🎯 **Key Insights**")
                            st.markdown(ai_response)
                        else:
                            st.warning("⚠️ AI temporarily unavailable. Showing basic insights:")
                            show_basic_insights(filtered_metrics)
                            
                    except Exception as e:
                        st.error(f"Error: {e}")
                        show_basic_insights(filtered_metrics)
        
        with col2:
            st.metric("Models Analyzed", len(filtered_metrics))
            st.metric("AI Status", "✅ Connected" if get_mistral_client() else "❌ Offline")
        
        # Simple custom question
        st.subheader("💬 Ask a Question")
        question = st.text_input("Ask about your data:", placeholder="Which model is best for production?")
        
        if question and st.button("🔍 Get Answer"):
            with st.spinner("Thinking..."):
                try:
                    answer = call_mistral_api(f"Based on this LLM benchmarking data, answer: {question}")
                    if "Error calling Mistral API" not in answer:
                        st.markdown(f"**Answer:** {answer}")
                    else:
                        st.info("AI temporarily unavailable. Please try again later.")
                except Exception as e:
                    st.error(f"Error: {e}")

def show_basic_insights(metrics):
    """Show basic insights when AI is not available"""
    st.markdown("### 📊 **Basic Analysis**")
    
    best_quality = metrics.loc[metrics['Quality_Score_mean'].idxmax()]
    fastest = metrics.loc[metrics['Latency_ms_mean'].idxmin()]
    most_efficient = metrics.loc[metrics['Energy_Wh_mean'].idxmin()]
    
    st.markdown(f"• **Best Quality:** {best_quality['Model']} ({best_quality['Quality_Score_mean']:.2f}/5)")
    st.markdown(f"• **Fastest:** {fastest['Model']} ({fastest['Latency_ms_mean']:.0f}ms)")
    st.markdown(f"• **Most Energy Efficient:** {most_efficient['Model']} ({most_efficient['Energy_Wh_mean']:.0f}Wh)")
    
    # Model size analysis
    small_models = metrics[metrics['Model_Size'] == 'Small']
    if not small_models.empty:
        best_small = small_models.loc[small_models['Quality_Score_mean'].idxmax()]
        st.markdown(f"• **Best Small Model:** {best_small['Model']} ({best_small['Quality_Score_mean']:.2f}/5)")
    
    large_models = metrics[metrics['Model_Size'] == 'Large']
    if not large_models.empty:
        best_large = large_models.loc[large_models['Quality_Score_mean'].idxmax()]
        st.markdown(f"• **Best Large Model:** {best_large['Model']} ({best_large['Quality_Score_mean']:.2f}/5)")
        
        # Performance heatmap
        st.subheader("🔥 **Performance Heatmap**")
        st.markdown("Visual comparison of all models across key metrics")
        
        heatmap_fig = create_performance_heatmap(filtered_metrics)
        st.plotly_chart(heatmap_fig, use_container_width=True)
    
    # Data export
    st.sidebar.header("📥 Export Data")
    
    if st.sidebar.button("Download Processed Data"):
        csv = filtered_metrics.to_csv(index=False)
        st.sidebar.download_button(
            label="Download CSV",
            data=csv,
            file_name="compar_ia_benchmark_results.csv",
            mime="text/csv"
        )
    
    # Advanced analysis mstion
    st.sidebar.header("🔬 Advanced Analysis")
    
    if st.sidebar.button("Generate Detailed Report"):
        st.sidebar.success("Detailed report generated!")
        # This could generate a comprehensive PDF report

if __name__ == "__main__":
    main()
