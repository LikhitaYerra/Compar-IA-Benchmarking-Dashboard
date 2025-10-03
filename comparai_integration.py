#!/usr/bin/env python3
"""
Integration script for ComparAI template with dashboard
"""

import pandas as pd
import openpyxl
from openpyxl import load_workbook
import streamlit as st

def load_comparai_data():
    """Load data from the ComparAI template"""
    try:
        # Load the Excel file
        wb = load_workbook('ComparAI_Benchmark_Template_v2-2.xlsx', data_only=True)
        
        # Find the main data sheet
        main_sheet = None
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            if ws.max_row > 1:
                main_sheet = ws
                break
        
        if not main_sheet:
            st.error("No data sheet found in the template")
            return None
        
        # Convert to DataFrame
        data = []
        headers = []
        
        # Get headers
        for col in range(1, main_sheet.max_column + 1):
            header = main_sheet.cell(row=1, column=col).value
            if header:
                headers.append(str(header))
            else:
                headers.append(f"Column_{col}")
        
        # Get data
        for row in range(2, main_sheet.max_row + 1):
            row_data = []
            has_data = False
            
            for col in range(1, main_sheet.max_column + 1):
                cell_value = main_sheet.cell(row=row, column=col).value
                row_data.append(cell_value)
                if cell_value is not None:
                    has_data = True
            
            if has_data:
                data.append(row_data)
        
        df = pd.DataFrame(data, columns=headers)
        
        # Map columns to dashboard format if needed
        column_mapping = {
            # Add mappings based on your template structure
            # 'Template Column': 'Dashboard Column'
        }
        
        # Apply mappings
        for old_col, new_col in column_mapping.items():
            if old_col in df.columns:
                df[new_col] = df[old_col]
        
        return df
        
    except Exception as e:
        st.error(f"Error loading ComparAI data: {e}")
        return None

# Add this function to your dashboard.py
# Replace the load_data() function with this one
