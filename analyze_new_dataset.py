#!/usr/bin/env python3
"""
Analyze the new complete ComparAI dataset v2-3
"""

import pandas as pd
import openpyxl
from openpyxl import load_workbook
import numpy as np

def analyze_new_dataset():
    """Analyze the new complete dataset"""
    try:
        # Load the new workbook
        wb = load_workbook('ComparAI_Benchmark_Template_v2-3.xlsx', data_only=True)
        
        print("🔍 Analyzing New ComparAI Dataset v2-3")
        print("=" * 50)
        
        # List all sheet names
        print(f"📋 Sheets found: {wb.sheetnames}")
        print()
        
        # Analyze each sheet
        for sheet_name in wb.sheetnames:
            print(f"📄 Sheet: {sheet_name}")
            print("-" * 30)
            
            ws = wb[sheet_name]
            
            # Get sheet dimensions
            max_row = ws.max_row
            max_col = ws.max_column
            print(f"   Dimensions: {max_row} rows × {max_col} columns")
            
            # Get header row
            headers = []
            for col in range(1, max_col + 1):
                cell_value = ws.cell(row=1, column=col).value
                if cell_value:
                    headers.append(str(cell_value))
                else:
                    headers.append(f"Column_{col}")
            
            print(f"   Headers: {headers}")
            
            # Count non-empty rows
            non_empty_rows = 0
            for row in range(2, max_row + 1):
                has_data = False
                for col in range(1, max_col + 1):
                    if ws.cell(row=row, column=col).value is not None:
                        has_data = True
                        break
                if has_data:
                    non_empty_rows += 1
            
            print(f"   Data rows: {non_empty_rows}")
            print()
        
        return wb
        
    except Exception as e:
        print(f"❌ Error analyzing file: {e}")
        return None

def convert_new_dataset_to_csv():
    """Convert the new dataset to CSV format"""
    try:
        wb = load_workbook('ComparAI_Benchmark_Template_v2-3.xlsx', data_only=True)
        
        # Find the main data sheet (usually 'Runs' or the largest sheet)
        main_sheet = None
        max_data_rows = 0
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            if ws.max_row > max_data_rows:
                max_data_rows = ws.max_row
                main_sheet = ws
        
        if not main_sheet:
            print("❌ No data sheet found")
            return None
        
        print(f"📊 Converting data from sheet: {main_sheet.title}")
        
        # Convert to DataFrame
        data = []
        headers = []
        
        # Get headers from first row
        for col in range(1, main_sheet.max_column + 1):
            header = main_sheet.cell(row=1, column=col).value
            if header:
                headers.append(str(header))
            else:
                headers.append(f"Column_{col}")
        
        # Get data rows
        for row in range(2, main_sheet.max_row + 1):
            row_data = []
            has_data = False
            
            for col in range(1, main_sheet.max_column + 1):
                cell_value = main_sheet.cell(row=row, column=col).value
                row_data.append(cell_value)
                if cell_value is not None:
                    has_data = True
            
            if has_data:
                data.append(row_data)
        
        # Create DataFrame
        df = pd.DataFrame(data, columns=headers)
        
        print(f"✅ Converted {len(df)} rows with columns: {list(df.columns)}")
        
        # Save as CSV
        df.to_csv('comparai_complete_dataset_v2-3.csv', index=False)
        print(f"💾 Saved complete dataset to: comparai_complete_dataset_v2-3.csv")
        
        # Show sample data
        print("\n📊 Sample of new data:")
        print(df.head())
        
        # Show data summary
        print(f"\n📈 Data summary:")
        print(f"   Total rows: {len(df)}")
        print(f"   Total columns: {len(df.columns)}")
        
        # Check for key columns
        key_columns = ['Model', 'Quality', 'Latency', 'Energy', 'co2']
        for col in key_columns:
            if col in df.columns:
                non_null_count = df[col].notna().sum()
                print(f"   {col}: {non_null_count} non-null values")
        
        return df
        
    except Exception as e:
        print(f"❌ Error converting data: {e}")
        return None

def update_dashboard_for_new_dataset():
    """Update the dashboard to use the new dataset"""
    try:
        # Read the dashboard file
        with open('dashboard_comparai.py', 'r') as f:
            dashboard_content = f.read()
        
        # Update the filename in the load function
        updated_content = dashboard_content.replace(
            'ComparAI_Benchmark_Template_v2-2.xlsx',
            'ComparAI_Benchmark_Template_v2-3.xlsx'
        )
        
        # Write the updated dashboard
        with open('dashboard_comparai.py', 'w') as f:
            f.write(updated_content)
        
        print("✅ Updated dashboard to use new dataset v2-3")
        
    except Exception as e:
        print(f"❌ Error updating dashboard: {e}")

def main():
    print("🚀 ComparAI Complete Dataset Analysis")
    print("=" * 40)
    
    # Analyze the new dataset
    wb = analyze_new_dataset()
    
    if wb:
        # Convert to CSV
        df = convert_new_dataset_to_csv()
        
        if df is not None:
            # Update dashboard
            update_dashboard_for_new_dataset()
            
            print("\n🎉 Analysis complete!")
            print("\nNext steps:")
            print("1. The dashboard has been updated to use the new dataset")
            print("2. Restart the dashboard to see the new data")
            print("3. All AI insights will now be based on the complete dataset")
        else:
            print("❌ Failed to convert data")
    else:
        print("❌ Failed to analyze dataset")

if __name__ == "__main__":
    main()
