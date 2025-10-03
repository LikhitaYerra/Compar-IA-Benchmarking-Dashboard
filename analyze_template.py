#!/usr/bin/env python3
"""
Analyze the ComparAI benchmark template and integrate with dashboard
"""

import pandas as pd
import openpyxl
from openpyxl import load_workbook
import sys

def analyze_excel_template(file_path):
    """Analyze the structure of the Excel template"""
    try:
        # Load the workbook
        wb = load_workbook(file_path, data_only=True)
        
        print(f"📊 Analyzing: {file_path}")
        print("=" * 50)
        
        # List all sheet names
        print(f"📋 Sheets found: {wb.sheetnames}")
        print()
        
        # Analyze each sheet
        for sheet_name in wb.sheetnames:
            print(f"📄 Sheet: {sheet_name}")
            print("-" * 30)
            
            ws = wb[sheet_name]
            
            # Get sheet dimensions
            max_row = ws.max_row
            max_col = ws.max_column
            print(f"   Dimensions: {max_row} rows × {max_col} columns")
            
            # Get header row (first row with data)
            headers = []
            for col in range(1, max_col + 1):
                cell_value = ws.cell(row=1, column=col).value
                if cell_value:
                    headers.append(str(cell_value))
                else:
                    headers.append(f"Column_{col}")
            
            print(f"   Headers: {headers}")
            
            # Count non-empty rows
            non_empty_rows = 0
            for row in range(2, max_row + 1):
                has_data = False
                for col in range(1, max_col + 1):
                    if ws.cell(row=row, column=col).value is not None:
                        has_data = True
                        break
                if has_data:
                    non_empty_rows += 1
            
            print(f"   Data rows: {non_empty_rows}")
            print()
        
        return wb
        
    except Exception as e:
        print(f"❌ Error analyzing file: {e}")
        return None

def convert_to_dashboard_format(wb, output_file="converted_data.csv"):
    """Convert the template to dashboard-compatible format"""
    try:
        # Find the main data sheet (usually the first one with data)
        main_sheet = None
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            if ws.max_row > 1:  # Has data beyond headers
                main_sheet = ws
                break
        
        if not main_sheet:
            print("❌ No data sheet found")
            return None
        
        print(f"📊 Converting data from sheet: {main_sheet.title}")
        
        # Convert to DataFrame
        data = []
        headers = []
        
        # Get headers from first row
        for col in range(1, main_sheet.max_column + 1):
            header = main_sheet.cell(row=1, column=col).value
            if header:
                headers.append(str(header))
            else:
                headers.append(f"Column_{col}")
        
        # Get data rows
        for row in range(2, main_sheet.max_row + 1):
            row_data = []
            has_data = False
            
            for col in range(1, main_sheet.max_column + 1):
                cell_value = main_sheet.cell(row=row, column=col).value
                row_data.append(cell_value)
                if cell_value is not None:
                    has_data = True
            
            if has_data:
                data.append(row_data)
        
        # Create DataFrame
        df = pd.DataFrame(data, columns=headers)
        
        print(f"✅ Converted {len(df)} rows with columns: {list(df.columns)}")
        
        # Save as CSV
        df.to_csv(output_file, index=False)
        print(f"💾 Saved converted data to: {output_file}")
        
        return df
        
    except Exception as e:
        print(f"❌ Error converting data: {e}")
        return None

def create_integration_script(template_file):
    """Create a script to integrate the template with the dashboard"""
    script_content = f'''#!/usr/bin/env python3
"""
Integration script for ComparAI template with dashboard
"""

import pandas as pd
import openpyxl
from openpyxl import load_workbook
import streamlit as st

def load_comparai_data():
    """Load data from the ComparAI template"""
    try:
        # Load the Excel file
        wb = load_workbook('{template_file}', data_only=True)
        
        # Find the main data sheet
        main_sheet = None
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            if ws.max_row > 1:
                main_sheet = ws
                break
        
        if not main_sheet:
            st.error("No data sheet found in the template")
            return None
        
        # Convert to DataFrame
        data = []
        headers = []
        
        # Get headers
        for col in range(1, main_sheet.max_column + 1):
            header = main_sheet.cell(row=1, column=col).value
            if header:
                headers.append(str(header))
            else:
                headers.append(f"Column_{{col}}")
        
        # Get data
        for row in range(2, main_sheet.max_row + 1):
            row_data = []
            has_data = False
            
            for col in range(1, main_sheet.max_column + 1):
                cell_value = main_sheet.cell(row=row, column=col).value
                row_data.append(cell_value)
                if cell_value is not None:
                    has_data = True
            
            if has_data:
                data.append(row_data)
        
        df = pd.DataFrame(data, columns=headers)
        
        # Map columns to dashboard format if needed
        column_mapping = {{
            # Add mappings based on your template structure
            # 'Template Column': 'Dashboard Column'
        }}
        
        # Apply mappings
        for old_col, new_col in column_mapping.items():
            if old_col in df.columns:
                df[new_col] = df[old_col]
        
        return df
        
    except Exception as e:
        st.error(f"Error loading ComparAI data: {{e}}")
        return None

# Add this function to your dashboard.py
# Replace the load_data() function with this one
'''
    
    with open('comparai_integration.py', 'w') as f:
        f.write(script_content)
    
    print("🔧 Created integration script: comparai_integration.py")

def main():
    template_file = "ComparAI_Benchmark_Template_v2-2.xlsx"
    
    print("🔍 Analyzing ComparAI Benchmark Template")
    print("=" * 50)
    
    # Analyze the template
    wb = analyze_excel_template(template_file)
    
    if wb:
        # Convert to dashboard format
        df = convert_to_dashboard_format(wb)
        
        if df is not None:
            print("\\n📊 Sample of converted data:")
            print(df.head())
            print(f"\\n📈 Data shape: {df.shape}")
            
            # Create integration script
            create_integration_script(template_file)
            
            print("\\n✅ Analysis complete!")
            print("\\nNext steps:")
            print("1. Review the converted data structure")
            print("2. Update column mappings in comparai_integration.py")
            print("3. Integrate with your dashboard")
        else:
            print("❌ Failed to convert data")
    else:
        print("❌ Failed to analyze template")

if __name__ == "__main__":
    main()
